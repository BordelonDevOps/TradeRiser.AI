"""Excel Export Manager for TradeRiser.AI
Comprehensive Excel export functionality matching individual repository capabilities
"""
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Union, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference
import xlsxwriter
import os
import logging

logging.basicConfig(
    filename='traderiser.log',
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s'
)

class ExcelExportManager:
    """Comprehensive Excel export manager for all TradeRiser modules"""
    
    def __init__(self, output_dir: str = "exports"):
        self.output_dir = output_dir
        self.ensure_output_directory()
        
        # Define styling
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        
    def ensure_output_directory(self):
        """Ensure output directory exists"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            
    def generate_filename(self, base_name: str, extension: str = ".xlsx") -> str:
        """Generate timestamped filename"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return os.path.join(self.output_dir, f"{base_name}_{timestamp}{extension}")
        
    def style_worksheet(self, worksheet, data_range: str):
        """Apply professional styling to worksheet"""
        # Style headers
        for cell in worksheet[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
            
        # Style data cells
        for row in worksheet[data_range]:
            for cell in row:
                cell.border = self.border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
                    
    def export_portfolio_analysis(self, portfolio_data: Dict, filename: Optional[str] = None) -> str:
        """Export comprehensive portfolio analysis to Excel"""
        if not filename:
            filename = self.generate_filename("portfolio_analysis")
            
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Portfolio Summary
            if 'summary' in portfolio_data:
                summary_df = pd.DataFrame([portfolio_data['summary']])
                summary_df.to_excel(writer, sheet_name='Portfolio Summary', index=False)
                
            # Holdings Details
            if 'holdings' in portfolio_data:
                holdings_df = pd.DataFrame(portfolio_data['holdings'])
                holdings_df.to_excel(writer, sheet_name='Holdings', index=False)
                
            # Performance Metrics
            if 'performance' in portfolio_data:
                perf_df = pd.DataFrame(portfolio_data['performance'])
                perf_df.to_excel(writer, sheet_name='Performance', index=False)
                
            # Risk Analysis
            if 'risk_analysis' in portfolio_data:
                risk_df = pd.DataFrame([portfolio_data['risk_analysis']])
                risk_df.to_excel(writer, sheet_name='Risk Analysis', index=False)
                
            # Recommendations
            if 'recommendations' in portfolio_data:
                rec_df = pd.DataFrame(portfolio_data['recommendations'])
                rec_df.to_excel(writer, sheet_name='Recommendations', index=False)
                
        # Apply styling
        self._apply_portfolio_styling(filename)
        
        logging.info(f"Portfolio analysis exported to {filename}")
        return filename
        
    def export_etf_analysis(self, etf_data: Dict, filename: Optional[str] = None) -> str:
        """Export ETF analysis matching ThePassiveInvestor capabilities"""
        if not filename:
            filename = self.generate_filename("etf_analysis")
            
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # ETF Overview
            if 'overview' in etf_data:
                overview_df = pd.DataFrame([etf_data['overview']])
                overview_df.to_excel(writer, sheet_name='ETF Overview', index=False)
                
            # Holdings Analysis
            if 'holdings' in etf_data:
                holdings_df = pd.DataFrame(etf_data['holdings'])
                holdings_df.to_excel(writer, sheet_name='Holdings', index=False)
                
            # Performance Comparison
            if 'performance_comparison' in etf_data:
                comp_df = pd.DataFrame(etf_data['performance_comparison'])
                comp_df.to_excel(writer, sheet_name='Performance Comparison', index=False)
                
            # Sector Allocation
            if 'sector_allocation' in etf_data:
                sector_df = pd.DataFrame(etf_data['sector_allocation'])
                sector_df.to_excel(writer, sheet_name='Sector Allocation', index=False)
                
            # Expense Analysis
            if 'expense_analysis' in etf_data:
                expense_df = pd.DataFrame(etf_data['expense_analysis'])
                expense_df.to_excel(writer, sheet_name='Expense Analysis', index=False)
                
        self._apply_etf_styling(filename)
        
        logging.info(f"ETF analysis exported to {filename}")
        return filename
        
    def export_financial_statements(self, financial_data: Dict, filename: Optional[str] = None) -> str:
        """Export financial statements matching FinanceToolkit capabilities"""
        if not filename:
            filename = self.generate_filename("financial_statements")
            
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Income Statement
            if 'income_statement' in financial_data:
                income_df = pd.DataFrame(financial_data['income_statement'])
                income_df.to_excel(writer, sheet_name='Income Statement', index=True)
                
            # Balance Sheet
            if 'balance_sheet' in financial_data:
                balance_df = pd.DataFrame(financial_data['balance_sheet'])
                balance_df.to_excel(writer, sheet_name='Balance Sheet', index=True)
                
            # Cash Flow Statement
            if 'cash_flow' in financial_data:
                cf_df = pd.DataFrame(financial_data['cash_flow'])
                cf_df.to_excel(writer, sheet_name='Cash Flow', index=True)
                
            # Financial Ratios
            if 'ratios' in financial_data:
                ratios_df = pd.DataFrame(financial_data['ratios'])
                ratios_df.to_excel(writer, sheet_name='Financial Ratios', index=True)
                
            # Valuation Metrics
            if 'valuation' in financial_data:
                val_df = pd.DataFrame(financial_data['valuation'])
                val_df.to_excel(writer, sheet_name='Valuation', index=True)
                
        self._apply_financial_styling(filename)
        
        logging.info(f"Financial statements exported to {filename}")
        return filename
        
    def export_crypto_analysis(self, crypto_data: Dict, filename: Optional[str] = None) -> str:
        """Export cryptocurrency analysis"""
        if not filename:
            filename = self.generate_filename("crypto_analysis")
            
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Crypto Overview
            if 'overview' in crypto_data:
                overview_df = pd.DataFrame([crypto_data['overview']])
                overview_df.to_excel(writer, sheet_name='Crypto Overview', index=False)
                
            # Price Analysis
            if 'price_analysis' in crypto_data:
                price_df = pd.DataFrame(crypto_data['price_analysis'])
                price_df.to_excel(writer, sheet_name='Price Analysis', index=False)
                
            # Technical Indicators
            if 'technical_indicators' in crypto_data:
                tech_df = pd.DataFrame(crypto_data['technical_indicators'])
                tech_df.to_excel(writer, sheet_name='Technical Analysis', index=False)
                
            # Market Sentiment
            if 'sentiment' in crypto_data:
                sentiment_df = pd.DataFrame(crypto_data['sentiment'])
                sentiment_df.to_excel(writer, sheet_name='Market Sentiment', index=False)
                
        self._apply_crypto_styling(filename)
        
        logging.info(f"Crypto analysis exported to {filename}")
        return filename
        
    def export_commodities_analysis(self, commodities_data: Dict, filename: Optional[str] = None) -> str:
        """Export commodities analysis"""
        if not filename:
            filename = self.generate_filename("commodities_analysis")
            
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Commodities Overview
            if 'overview' in commodities_data:
                overview_df = pd.DataFrame(commodities_data['overview'])
                overview_df.to_excel(writer, sheet_name='Commodities Overview', index=False)
                
            # Price Trends
            if 'price_trends' in commodities_data:
                trends_df = pd.DataFrame(commodities_data['price_trends'])
                trends_df.to_excel(writer, sheet_name='Price Trends', index=False)
                
            # Supply & Demand
            if 'supply_demand' in commodities_data:
                sd_df = pd.DataFrame(commodities_data['supply_demand'])
                sd_df.to_excel(writer, sheet_name='Supply & Demand', index=False)
                
        self._apply_commodities_styling(filename)
        
        logging.info(f"Commodities analysis exported to {filename}")
        return filename
        
    def export_backtest_results(self, backtest_data: Dict, filename: Optional[str] = None) -> str:
        """Export backtest results"""
        if not filename:
            filename = self.generate_filename("backtest_results")
            
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Strategy Performance
            if 'performance' in backtest_data:
                perf_df = pd.DataFrame(backtest_data['performance'])
                perf_df.to_excel(writer, sheet_name='Strategy Performance', index=False)
                
            # Trade History
            if 'trades' in backtest_data:
                trades_df = pd.DataFrame(backtest_data['trades'])
                trades_df.to_excel(writer, sheet_name='Trade History', index=False)
                
            # Risk Metrics
            if 'risk_metrics' in backtest_data:
                risk_df = pd.DataFrame([backtest_data['risk_metrics']])
                risk_df.to_excel(writer, sheet_name='Risk Metrics', index=False)
                
            # Drawdown Analysis
            if 'drawdown' in backtest_data:
                dd_df = pd.DataFrame(backtest_data['drawdown'])
                dd_df.to_excel(writer, sheet_name='Drawdown Analysis', index=False)
                
        self._apply_backtest_styling(filename)
        
        logging.info(f"Backtest results exported to {filename}")
        return filename
        
    def export_alpaca_trading_data(self, trading_data: Dict, filename: Optional[str] = None) -> str:
        """Export Alpaca trading data"""
        if not filename:
            filename = self.generate_filename("alpaca_trading_data")
            
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Account Summary
            if 'account' in trading_data:
                account_df = pd.DataFrame([trading_data['account']])
                account_df.to_excel(writer, sheet_name='Account Summary', index=False)
                
            # Positions
            if 'positions' in trading_data:
                positions_df = pd.DataFrame(trading_data['positions'])
                positions_df.to_excel(writer, sheet_name='Positions', index=False)
                
            # Orders
            if 'orders' in trading_data:
                orders_df = pd.DataFrame(trading_data['orders'])
                orders_df.to_excel(writer, sheet_name='Orders', index=False)
                
            # Performance
            if 'performance' in trading_data:
                perf_df = pd.DataFrame(trading_data['performance'])
                perf_df.to_excel(writer, sheet_name='Performance', index=False)
                
        self._apply_trading_styling(filename)
        
        logging.info(f"Alpaca trading data exported to {filename}")
        return filename
        
    def export_comprehensive_report(self, all_data: Dict, filename: Optional[str] = None) -> str:
        """Export comprehensive report with all modules data"""
        if not filename:
            filename = self.generate_filename("comprehensive_report")
            
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Executive Summary
            if 'executive_summary' in all_data:
                summary_df = pd.DataFrame([all_data['executive_summary']])
                summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
                
            # Portfolio Data
            if 'portfolio' in all_data:
                for key, data in all_data['portfolio'].items():
                    if isinstance(data, (list, dict)):
                        df = pd.DataFrame(data) if isinstance(data, list) else pd.DataFrame([data])
                        sheet_name = f"Portfolio_{key.replace(' ', '_')}"
                        df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                        
            # ETF Data
            if 'etf' in all_data:
                for key, data in all_data['etf'].items():
                    if isinstance(data, (list, dict)):
                        df = pd.DataFrame(data) if isinstance(data, list) else pd.DataFrame([data])
                        sheet_name = f"ETF_{key.replace(' ', '_')}"
                        df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                        
            # Crypto Data
            if 'crypto' in all_data:
                for key, data in all_data['crypto'].items():
                    if isinstance(data, (list, dict)):
                        df = pd.DataFrame(data) if isinstance(data, list) else pd.DataFrame([data])
                        sheet_name = f"Crypto_{key.replace(' ', '_')}"
                        df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                        
            # Commodities Data
            if 'commodities' in all_data:
                for key, data in all_data['commodities'].items():
                    if isinstance(data, (list, dict)):
                        df = pd.DataFrame(data) if isinstance(data, list) else pd.DataFrame([data])
                        sheet_name = f"Commodities_{key.replace(' ', '_')}"
                        df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                        
        self._apply_comprehensive_styling(filename)
        
        logging.info(f"Comprehensive report exported to {filename}")
        return filename
        
    def _apply_portfolio_styling(self, filename: str):
        """Apply styling to portfolio analysis workbook"""
        workbook = openpyxl.load_workbook(filename)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            self._style_worksheet_basic(worksheet)
        workbook.save(filename)
        
    def _apply_etf_styling(self, filename: str):
        """Apply styling to ETF analysis workbook"""
        workbook = openpyxl.load_workbook(filename)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            self._style_worksheet_basic(worksheet)
        workbook.save(filename)
        
    def _apply_financial_styling(self, filename: str):
        """Apply styling to financial statements workbook"""
        workbook = openpyxl.load_workbook(filename)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            self._style_worksheet_basic(worksheet)
        workbook.save(filename)
        
    def _apply_crypto_styling(self, filename: str):
        """Apply styling to crypto analysis workbook"""
        workbook = openpyxl.load_workbook(filename)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            self._style_worksheet_basic(worksheet)
        workbook.save(filename)
        
    def _apply_commodities_styling(self, filename: str):
        """Apply styling to commodities analysis workbook"""
        workbook = openpyxl.load_workbook(filename)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            self._style_worksheet_basic(worksheet)
        workbook.save(filename)
        
    def _apply_backtest_styling(self, filename: str):
        """Apply styling to backtest results workbook"""
        workbook = openpyxl.load_workbook(filename)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            self._style_worksheet_basic(worksheet)
        workbook.save(filename)
        
    def _apply_trading_styling(self, filename: str):
        """Apply styling to trading data workbook"""
        workbook = openpyxl.load_workbook(filename)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            self._style_worksheet_basic(worksheet)
        workbook.save(filename)
        
    def _apply_comprehensive_styling(self, filename: str):
        """Apply styling to comprehensive report workbook"""
        workbook = openpyxl.load_workbook(filename)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            self._style_worksheet_basic(worksheet)
        workbook.save(filename)
        
    def _style_worksheet_basic(self, worksheet):
        """Apply basic styling to worksheet"""
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
        # Style headers if they exist
        if worksheet.max_row > 0:
            for cell in worksheet[1]:
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.border
                
    def get_export_summary(self) -> Dict:
        """Get summary of available export capabilities"""
        return {
            'available_exports': [
                'Portfolio Analysis',
                'ETF Analysis',
                'Financial Statements',
                'Crypto Analysis',
                'Commodities Analysis',
                'Backtest Results',
                'Alpaca Trading Data',
                'Comprehensive Report'
            ],
            'supported_formats': ['Excel (.xlsx)', 'CSV'],
            'features': [
                'Professional styling',
                'Multiple worksheets',
                'Auto-sized columns',
                'Timestamped filenames',
                'Comprehensive data coverage'
            ],
            'output_directory': self.output_dir
        }